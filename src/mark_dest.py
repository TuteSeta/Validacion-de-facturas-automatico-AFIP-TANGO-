# src/mark_dest.py
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
from pathlib import Path
import pandas as pd
import numpy as np

from src.transform import _normalize_cuit, _normalize_name, _to_number_locale

YELLOW = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")  # diferencias

def _coerce_for_compare(x, kind):
    if kind == "number":
        v = _to_number_locale(x)
        return np.nan if v is pd.NA else float(v)
    if kind == "date":
        try:
            return pd.to_datetime(x).date()
        except Exception:
            return np.nan
    # string
    s = "" if x is pd.NA else str(x)
    return s.strip().upper()

def _compare_with_tolerance(a, b, kind, tol):
    a = _coerce_for_compare(a, kind)
    b = _coerce_for_compare(b, kind)
    if kind == "number":
        if (isinstance(a, float) and np.isnan(a)) and (isinstance(b, float) and np.isnan(b)):
            return True
        if isinstance(a, float) and isinstance(b, float):
            return abs(a - b) <= tol
        return False
    else:
        a = "" if (isinstance(a, float) and np.isnan(a)) else a
        b = "" if (isinstance(b, float) and np.isnan(b)) else b
        return a == b

def _build_ws_key_index(ws, colmap):
    """
    Recorre la hoja destino para construir un índice:
    key = (N_COMP_normalizado, CUIT_normalizado) -> lista de rownums
    colmap: dict nombre_col -> indice_columna (1-based)
    """
    key_to_rows = {}
    for r in range(2, ws.max_row + 1):
        ncomp_cell = ws.cell(row=r, column=colmap["N_COMP"]).value
        cuit_cell  = ws.cell(row=r, column=colmap["IDENTIFTRI"]).value
        ncomp = str(ncomp_cell).strip().upper() if ncomp_cell is not None else ""
        cuit_norm = _normalize_cuit(cuit_cell)
        cuit = "" if pd.isna(cuit_norm) else str(cuit_norm)   # ← evita "or ''" con pd.NA
        key = (ncomp, cuit)
        key_to_rows.setdefault(key, []).append(r)
    return key_to_rows

def _ensure_headers(ws, needed):
    """Devuelve un dict nombre_col -> idx, error si falta alguna columna necesaria."""
    header = {ws.cell(row=1, column=j).value: j for j in range(1, ws.max_column + 1)}
    missing = [c for c in needed if c not in header]
    if missing:
        raise KeyError(f"En la hoja '{ws.title}' faltan columnas: {missing}")
    return header

def mark_and_append(
    origen_df: pd.DataFrame,
    destino_xlsx_path: str,
    destino_sheet: str,
    columns_cfg: list,
    out_path: str,
):
    """
    - Abre el Excel de destino desde disco (sin copiar con shutil) y lo guarda como un archivo nuevo.
    - Marca en amarillo las celdas de Tango que no coinciden contra AFIP.
    - NO inserta filas nuevas; solo devuelve cuántas faltaron (missing_count).
    """
    # 1) Abrimos el workbook de destino original
    try:
        wb = load_workbook(destino_xlsx_path)
    except PermissionError:
        raise PermissionError(
            f"No se pudo abrir '{destino_xlsx_path}'. Cerrá el archivo si está abierto en Excel."
        )

    out_file = Path(out_path)
    out_file.parent.mkdir(parents=True, exist_ok=True)

    def save_wb_safely(workbook, path: Path):
        try:
            workbook.save(path)
            return path
        except PermissionError:
            from datetime import datetime
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            alt = path.with_name(f"{path.stem}_{ts}{path.suffix}")
            workbook.save(alt)
            return alt

    # 2) Seleccionamos la hoja del workbook cargado
    if destino_sheet not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{destino_sheet}' en {destino_xlsx_path}")
    ws = wb[destino_sheet]

    # 3) Chequear encabezados necesarios
    needed = set(["N_COMP", "IDENTIFTRI"]) | {c["name"] for c in columns_cfg}
    header = _ensure_headers(ws, needed)

    # 4) Índice por clave en la hoja: (N_COMP, IDENTIFTRI) -> filas
    key_to_rows = _build_ws_key_index(ws, header)

    # 5) Recorremos AFIP y validamos
    missing_afip_rows = []

    for _, row in origen_df.iterrows():
        ncomp = str(row["N_COMP"]).strip().upper()
        cuit_norm = _normalize_cuit(row["IDENTIFTRI"])
        cuit = "" if pd.isna(cuit_norm) else str(cuit_norm)
        key = (ncomp, cuit)

        if key in key_to_rows and len(key_to_rows[key]) > 0:
            r = key_to_rows[key][0]
            letter = ncomp[0] if ncomp else ""
            if letter == "C":
                cols_to_check = [c for c in columns_cfg if c["name"] == "IMP_TOTAL"]
            else:
                cols_to_check = columns_cfg
            tc_raw = row.get("TC", 1.0)
            tc = _to_number_locale(tc_raw)
            if not isinstance(tc, float) or np.isnan(tc):
                tc = 1.0
            for colinfo in cols_to_check:
                name = colinfo["name"]
                kind = colinfo.get("type", "string")
                tol  = float(colinfo.get("tolerance", 0.0))
                ws_val = ws.cell(row=r, column=header[name]).value
                if kind == "number":
                    a_num = _to_number_locale(row[name])
                    a_adj = (a_num * tc) if (isinstance(a_num, float) and not np.isnan(a_num)) else a_num
                    ok = _compare_with_tolerance(a_adj, ws_val, kind, tol)
                else:
                    ok = _compare_with_tolerance(row[name], ws_val, kind, tol)

                if not ok:
                    ws.cell(row=r, column=header[name]).fill = YELLOW
                    current_font = ws.cell(row=r, column=header[name]).font
                    ws.cell(row=r, column=header[name]).font = Font(
                        name=getattr(current_font, "name", None),
                        size=getattr(current_font, "sz", None),
                        bold=True,
                        underline="single",
                    )
        else:
            missing_afip_rows.append(row)  # solo contamos, no insertamos

    # 6) Guardar de forma segura
    missing_count = len(missing_afip_rows)
    save_wb_safely(wb, out_file)
    return missing_count
