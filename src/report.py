from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import pandas as pd

YELLOW = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

def write_report(path_output: str, df: pd.DataFrame, keys, columns_cfg):
    """
    Genera un Excel con las diferencias marcadas en amarillo.
    """
    # Exportamos primero a Excel plano
    cols_order = keys[:]
    for c in columns_cfg:
        name = c["name"]
        cols_order += [f"{name}_origen", f"{name}_destino"]
    cols_order += ["__row_ok__"]

    df_export = df[cols_order].copy()
    df_export.to_excel(path_output, sheet_name="Validación", index=False)

    # Reabrimos para pintar diferencias
    wb = load_workbook(path_output)
    ws = wb["Validación"]

    # Mapeo de encabezados (columna -> índice)
    header = {ws.cell(row=1, column=j).value: j for j in range(1, ws.max_column + 1)}

    for i in range(2, ws.max_row + 1):
        for c in columns_cfg:
            name = c["name"]
            ok = df.iloc[i - 2][f"__match__{name}"]
            if not ok:
                j1 = header.get(f"{name}_origen")
                j2 = header.get(f"{name}_destino")
                if j1:
                    ws.cell(row=i, column=j1).fill = YELLOW
                if j2:
                    ws.cell(row=i, column=j2).fill = YELLOW

    wb.save(path_output)
