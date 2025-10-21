import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

from src.transform import _resolve_col, _tipo_to_letter, _to_int_safe, _normalize_cuit, _to_number_locale

GREEN_FILL  = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

def _build_ncomp_from_parts_row(row, df, amap):
    c_tipo = _resolve_col(df, amap["tipo"]) if "tipo" in amap else None
    c_pv   = _resolve_col(df, amap["pv"])   if "pv"   in amap else None
    c_num  = _resolve_col(df, amap["num"])  if "num"  in amap else None
    pattern = amap.get("build_pattern", "{letter}{pv:04d}{num:08d}")

    letter = _tipo_to_letter(row.get(c_tipo)) if c_tipo else "A"
    pv_i   = _to_int_safe(row.get(c_pv)) if c_pv else 0
    num_i  = _to_int_safe(row.get(c_num)) if c_num else 0
    return pattern.format(letter=letter, pv=pv_i, num=num_i)

def _compute_status_series(full_df: pd.DataFrame, destino_df: pd.DataFrame, amap: dict, tolerances: dict) -> pd.Series:
    work = full_df.copy()
    work["N_COMP"] = work.apply(lambda r: _build_ncomp_from_parts_row(r, full_df, amap), axis=1).astype(str).str.strip().str.upper()

    c_cuit = _resolve_col(full_df, amap.get("cuit")) if amap.get("cuit") else None
    work["IDENTIFTRI"] = work[c_cuit].map(_normalize_cuit) if c_cuit in full_df.columns else pd.NA

    c_tc = _resolve_col(full_df, amap.get("exchange_rate")) if amap.get("exchange_rate") else None
    if c_tc and c_tc in full_df.columns:
        work["TC"] = work[c_tc].map(_to_number_locale).fillna(1.0)
    else:
        work["TC"] = 1.0
        
    mi = amap.get("importes", {})
    def take_num(colkey):
        ck = _resolve_col(full_df, mi.get(colkey))
        return work[ck].map(_to_number_locale) if ck in full_df.columns else pd.NA

    work["IMP_EXENTO"] = take_num("exento")
    work["IMP_NETO"]   = take_num("neto")
    work["IMP_IVA"]    = take_num("iva")
    work["IMP_TOTAL"]  = take_num("total")

    keys = ["N_COMP", "IDENTIFTRI"]
    merged = work.merge(destino_df, on=keys, how="left", suffixes=("_origen", "_destino"), indicator=True)

    status = []
    for _, row in merged.iterrows():
        if row.get("_merge") == "left_only":
            status.append("Omitida")
            continue
        ncomp = str(row.get("N_COMP", "")).strip().upper()
        letter = ncomp[0] if ncomp else ""
        cols_to_check = ("IMP_TOTAL",) if letter == "C" else ("IMP_EXENTO", "IMP_NETO", "IMP_IVA", "IMP_TOTAL")
        tc_raw = row.get("TC_origen", row.get("TC", 1.0))
        tc = _to_number_locale(tc_raw)
        if not isinstance(tc, float) or pd.isna(tc):
            tc = 1.0
            
        ok_all = True
        for name in cols_to_check:
            tol = float(tolerances.get(name, 0.0))
            av = row.get(f"{name}_origen")
            bv = row.get(f"{name}_destino")
            a = _to_number_locale(av)
            b = _to_number_locale(bv)
            a_adj = a * tc if (isinstance(a, float) and not pd.isna(a)) else a
            # Ambos NaN => ok
            if (isinstance(a_adj, float) and pd.isna(a_adj)) and (isinstance(b, float) and pd.isna(b)):
                continue
            if not (isinstance(a_adj, float) and isinstance(b, float) and abs(a_adj - b) <= tol):
                ok_all = False
                break
        status.append("Coincide" if ok_all else "No coincide")

    return pd.Series(status, index=work.index)

def write_origen_validado(origen_path: str, sheet: str, mapping: dict, destino_df: pd.DataFrame, tolerances: dict, out_path: str):
    # Leer origen completo preservando columnas y orden; encabezados reales en la segunda fila (header=1)
    full_df = pd.read_excel(origen_path, sheet_name=sheet, header=1)
    estados = _compute_status_series(full_df, destino_df, mapping["afip"], tolerances)

    export_df = full_df.copy()
    export_df["Estado_Validación"] = estados

    # Exportar a Excel
    export_df.to_excel(out_path, sheet_name="Origen", index=False)

    # Aplicar colores por fila
    wb = load_workbook(out_path)
    ws = wb["Origen"]

    # Columna del estado
    header = {ws.cell(row=1, column=j).value: j for j in range(1, ws.max_column + 1)}
    j_estado = header.get("Estado_Validación")

    for i in range(2, ws.max_row + 1):
        estado = ws.cell(row=i, column=j_estado).value
        if estado == "Coincide":
            fill = GREEN_FILL
        elif estado == "No coincide":
            fill = RED_FILL
        else:
            fill = YELLOW_FILL
        for j in range(1, ws.max_column + 1):
            ws.cell(row=i, column=j).fill = fill

    wb.save(out_path)


